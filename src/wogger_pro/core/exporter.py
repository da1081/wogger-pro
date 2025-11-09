"""Utilities for generating configurable time-based exports."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Iterable, Sequence

from .models import Entry


UNCATEGORIZED_LABEL = "uncategorized-entries"
CATEGORY_SEPARATOR = " - "


class ExportType(Enum):
    CATEGORIES = "categories"
    TASKS = "tasks"
    ENTRIES = "entries"


class TimeGrouping(Enum):
    HOURS = "hours"
    DAYS = "days"
    WEEKS = "weeks"
    MONTHS = "months"
    YEARS = "years"


class ExportFormat(Enum):
    CSV = "csv"
    JSONL = "jsonl"
    JSON = "json"
    EXCEL = "excel"


@dataclass(frozen=True)
class ExportOptions:
    start: datetime
    end: datetime
    export_type: ExportType
    grouping: TimeGrouping
    format: ExportFormat


@dataclass
class ExportTable:
    columns: list[str]
    rows: list[dict[str, object]]


def generate_export_table(entries: Sequence[Entry], options: ExportOptions) -> ExportTable:
    if options.start >= options.end:
        raise ValueError("Start must be before end")

    filtered = [
        entry
        for entry in entries
        if entry.segment_start < options.end and entry.segment_end > options.start
    ]

    if options.export_type is ExportType.CATEGORIES:
        return _generate_category_table(filtered, options)
    if options.export_type is ExportType.TASKS:
        return _generate_task_table(filtered, options)
    if options.export_type is ExportType.ENTRIES:
        return _generate_entries_table(filtered, options)
    raise ValueError(f"Unsupported export type: {options.export_type}")


def write_export(table: ExportTable, options: ExportOptions, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if options.format is ExportFormat.CSV:
        _write_csv(table, path)
        return
    if options.format is ExportFormat.JSON:
        _write_json(table, options, path)
        return
    if options.format is ExportFormat.JSONL:
        _write_jsonl(table, path)
        return
    if options.format is ExportFormat.EXCEL:
        _write_excel(table, options, path)
        return
    raise ValueError(f"Unsupported export format: {options.format}")


def create_jf_excel_export(entries: Sequence[Entry], categories: Sequence[str], path: Path) -> None:
    """Create a preconfigured JF Excel workbook grouped by category tree and date."""

    if not entries:
        raise ValueError("No entries available to export")

    ordering = _CategoryOrdering.from_categories(categories)
    normalized_categories = _collect_category_names(categories, entries, ordering)
    minutes_lookup = _build_minutes_lookup(entries)
    date_buckets = _collect_date_buckets(entries)

    tree_root = _build_category_tree(normalized_categories, ordering)
    _ensure_uncategorized_node(tree_root, minutes_lookup, ordering)
    _assign_minutes(tree_root, minutes_lookup)

    _write_jf_excel_workbook(tree_root, date_buckets, path)


# ---------------------------------------------------------------------------
# Export table builders

def _generate_category_table(entries: Sequence[Entry], options: ExportOptions) -> ExportTable:
    buckets = list(_iter_buckets(options.start, options.end, options.grouping))
    columns = ["category"] + [label for label, _start, _end in buckets]
    totals: dict[str, list[int]] = {}

    for entry in entries:
        label = (entry.category or UNCATEGORIZED_LABEL).strip() or UNCATEGORIZED_LABEL
        bucket_minutes = totals.setdefault(label, [0] * len(buckets))
        for index, (_label, bucket_start, bucket_end) in enumerate(buckets):
            minutes = _overlap_minutes(entry.segment_start, entry.segment_end, bucket_start, bucket_end)
            if minutes:
                bucket_minutes[index] += minutes

    rows: list[dict[str, object]] = []
    for category in sorted(totals):
        bucket_values = totals[category]
        row = {"category": category}
        for idx, (label, _start, _end) in enumerate(buckets):
            row[label] = bucket_values[idx]
        rows.append(row)

    return ExportTable(columns=columns, rows=rows)


def _generate_task_table(entries: Sequence[Entry], options: ExportOptions) -> ExportTable:
    buckets = list(_iter_buckets(options.start, options.end, options.grouping))
    columns = ["task"] + [label for label, _start, _end in buckets]
    totals: dict[str, list[int]] = {}

    for entry in entries:
        label = entry.task.strip()
        if not label:
            continue
        bucket_minutes = totals.setdefault(label, [0] * len(buckets))
        for index, (_label, bucket_start, bucket_end) in enumerate(buckets):
            minutes = _overlap_minutes(entry.segment_start, entry.segment_end, bucket_start, bucket_end)
            if minutes:
                bucket_minutes[index] += minutes

    rows: list[dict[str, object]] = []
    for task in sorted(totals):
        bucket_values = totals[task]
        row = {"task": task}
        for idx, (label, _start, _end) in enumerate(buckets):
            row[label] = bucket_values[idx]
        rows.append(row)

    return ExportTable(columns=columns, rows=rows)


def _generate_entries_table(entries: Sequence[Entry], options: ExportOptions) -> ExportTable:
    columns = [
        "entry_id",
        "task",
        "category",
        "minutes",
        "weekday",
        "group",
        "start",
        "end",
    ]

    rows: list[dict[str, object]] = []
    for entry in sorted(entries, key=lambda e: (e.segment_start, e.segment_end, e.entry_id)):
        group_label = _label_for_bucket(_floor_to_group(entry.segment_start, options.grouping), options.grouping)
        row = {
            "entry_id": entry.entry_id,
            "task": entry.task,
            "category": entry.category or "",
            "minutes": entry.minutes,
            "weekday": entry.segment_start.strftime("%A"),
            "group": group_label,
            "start": entry.segment_start.isoformat(timespec="seconds"),
            "end": entry.segment_end.isoformat(timespec="seconds"),
        }
        rows.append(row)

    return ExportTable(columns=columns, rows=rows)


# ---------------------------------------------------------------------------
# Writers

def _write_csv(table: ExportTable, path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=table.columns)
        writer.writeheader()
        for row in table.rows:
            writer.writerow({column: row.get(column, "") for column in table.columns})


def _write_json(table: ExportTable, options: ExportOptions, path: Path) -> None:
    payload = {
        "type": options.export_type.value,
        "grouping": options.grouping.value,
        "start": options.start.isoformat(timespec="seconds"),
        "end": options.end.isoformat(timespec="seconds"),
        "columns": table.columns,
        "rows": table.rows,
    }
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def _write_jsonl(table: ExportTable, path: Path) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in table.rows:
            handle.write(json.dumps(row, ensure_ascii=False))
            handle.write("\n")


def _write_excel(table: ExportTable, options: ExportOptions, path: Path) -> None:
    from openpyxl import Workbook  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    sheet.append(table.columns)
    for row in table.rows:
        sheet.append([row.get(column, "") for column in table.columns])

    for index, column_name in enumerate(table.columns, start=1):
        max_length = len(str(column_name))
        for row in table.rows:
            value = row.get(column_name, "")
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[get_column_letter(index)].width = max(10, min(max_length + 2, 60))

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["Type", options.export_type.value])
    metadata.append(["Grouping", options.grouping.value])
    metadata.append(["Start", options.start.isoformat(timespec="seconds")])
    metadata.append(["End", options.end.isoformat(timespec="seconds")])
    metadata.append(["Rows", len(table.rows)])

    workbook.save(path)


# ---------------------------------------------------------------------------
# Time grouping helpers

def _iter_buckets(start: datetime, end: datetime, grouping: TimeGrouping) -> Iterable[tuple[str, datetime, datetime]]:
    cursor = _floor_to_group(start, grouping)
    while cursor < end:
        next_cursor = _advance_group(cursor, grouping)
        bucket_start = max(cursor, start)
        bucket_end = min(next_cursor, end)
        if bucket_start < bucket_end:
            yield _label_for_bucket(cursor, grouping), bucket_start, bucket_end
        cursor = next_cursor


def _floor_to_group(moment: datetime, grouping: TimeGrouping) -> datetime:
    if grouping is TimeGrouping.HOURS:
        return moment.replace(minute=0, second=0, microsecond=0)
    if grouping is TimeGrouping.DAYS:
        return moment.replace(hour=0, minute=0, second=0, microsecond=0)
    if grouping is TimeGrouping.WEEKS:
        base = moment.replace(hour=0, minute=0, second=0, microsecond=0)
        return base - timedelta(days=base.weekday())
    if grouping is TimeGrouping.MONTHS:
        return moment.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if grouping is TimeGrouping.YEARS:
        return moment.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    raise ValueError(f"Unsupported time grouping: {grouping}")


def _advance_group(moment: datetime, grouping: TimeGrouping) -> datetime:
    if grouping is TimeGrouping.HOURS:
        return moment + timedelta(hours=1)
    if grouping is TimeGrouping.DAYS:
        return moment + timedelta(days=1)
    if grouping is TimeGrouping.WEEKS:
        return moment + timedelta(weeks=1)
    if grouping is TimeGrouping.MONTHS:
        year = moment.year
        month = moment.month + 1
        if month == 13:
            month = 1
            year += 1
        return moment.replace(year=year, month=month, day=1)
    if grouping is TimeGrouping.YEARS:
        return moment.replace(year=moment.year + 1, month=1, day=1)
    raise ValueError(f"Unsupported time grouping: {grouping}")


def _label_for_bucket(moment: datetime, grouping: TimeGrouping) -> str:
    if grouping is TimeGrouping.HOURS:
        return moment.strftime("%Y-%m-%d %H:00")
    if grouping is TimeGrouping.DAYS:
        return moment.strftime("%Y-%m-%d")
    if grouping is TimeGrouping.WEEKS:
        iso_year, iso_week, _ = moment.isocalendar()
        return f"{iso_year}-W{iso_week:02d}"
    if grouping is TimeGrouping.MONTHS:
        return moment.strftime("%Y-%m")
    if grouping is TimeGrouping.YEARS:
        return moment.strftime("%Y")
    raise ValueError(f"Unsupported time grouping: {grouping}")


def _overlap_minutes(start_a: datetime, end_a: datetime, start_b: datetime, end_b: datetime) -> int:
    latest_start = max(start_a, start_b)
    earliest_end = min(end_a, end_b)
    if earliest_end <= latest_start:
        return 0
    delta = earliest_end - latest_start
    return max(0, int(delta.total_seconds() // 60))


# ---------------------------------------------------------------------------
# JF Excel helpers


@dataclass
class _CategoryNode:
    name: str
    full_name: str | None
    depth: int
    children: list["_CategoryNode"] = field(default_factory=list)
    minutes: dict[date, int] = field(default_factory=dict)


@dataclass
class _CategoryOrdering:
    index_by_path: dict[tuple[str, ...], int]
    next_index: int

    @classmethod
    def from_categories(cls, categories: Sequence[str]) -> "_CategoryOrdering":
        index_map: dict[tuple[str, ...], int] = {}
        next_index = 0
        for category in categories:
            parts = _split_category_path(category)
            if not parts:
                continue
            path: list[str] = []
            for part in parts:
                path.append(part)
                key = tuple(segment.lower() for segment in path)
                if key not in index_map:
                    index_map[key] = next_index
                    next_index += 1
        return cls(index_by_path=index_map, next_index=next_index)

    def ensure(self, parts: Sequence[str]) -> int:
        key = tuple(part.lower() for part in parts if part)
        if not key:
            return -1
        existing = self.index_by_path.get(key)
        if existing is not None:
            return existing
        index = self.next_index
        self.index_by_path[key] = index
        self.next_index += 1
        return index

    def order_for(self, parts: Sequence[str]) -> int:
        return self.ensure(parts)


def _split_category_path(category: str) -> list[str]:
    return [segment.strip() for segment in category.split(CATEGORY_SEPARATOR) if segment.strip()]


def _collect_category_names(
    categories: Sequence[str],
    entries: Sequence[Entry],
    ordering: _CategoryOrdering,
) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()

    for value in categories:
        normalized = _normalize_category(value)
        if not normalized:
            continue
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        ordering.ensure(_split_category_path(normalized))
        ordered.append(normalized)

    extras: set[str] = set()
    for entry in entries:
        normalized = _normalize_category(entry.category)
        if not normalized:
            continue
        key = normalized.lower()
        if key in seen:
            continue
        extras.add(normalized)

    for value in sorted(extras, key=lambda item: item.lower()):
        ordering.ensure(_split_category_path(value))
        ordered.append(value)
        seen.add(value.lower())

    return ordered


def _normalize_category(value: str | None) -> str:
    return (value or "").strip()


def _build_category_tree(categories: Sequence[str], ordering: _CategoryOrdering) -> _CategoryNode:
    root = _CategoryNode(name="(root)", full_name=None, depth=0)
    for category in categories:
        parts = _split_category_path(category)
        if not parts:
            continue
        parent = root
        path: list[str] = []
        for part in parts:
            path.append(part)
            full_name = CATEGORY_SEPARATOR.join(path)
            ordering.ensure(path)
            child = _find_child(parent, part)
            if child is None:
                child = _CategoryNode(name=part, full_name=full_name, depth=parent.depth + 1)
                parent.children.append(child)
            parent = child
    _apply_ordering(root, ordering)
    return root


def _find_child(node: _CategoryNode, name: str) -> _CategoryNode | None:
    target = name.lower()
    for child in node.children:
        if child.name.lower() == target:
            return child
    return None


def _build_minutes_lookup(entries: Sequence[Entry]) -> dict[str, dict[date, int]]:
    minutes_lookup: dict[str, dict[date, int]] = {}
    for entry in entries:
        category = _normalize_category(entry.category) or UNCATEGORIZED_LABEL
        key = category.lower()
        bucket = minutes_lookup.setdefault(key, {})
        day = entry.segment_start.date()
        bucket[day] = bucket.get(day, 0) + entry.minutes
    return minutes_lookup


def _ensure_uncategorized_node(
    root: _CategoryNode,
    minutes_lookup: dict[str, dict[date, int]],
    ordering: _CategoryOrdering,
) -> None:
    if UNCATEGORIZED_LABEL.lower() not in minutes_lookup:
        return
    existing = _find_child(root, UNCATEGORIZED_LABEL)
    if existing is None:
        uncategorized = _CategoryNode(
            name=UNCATEGORIZED_LABEL,
            full_name=UNCATEGORIZED_LABEL,
            depth=1,
        )
        root.children.append(uncategorized)
        ordering.ensure([UNCATEGORIZED_LABEL])
        _apply_ordering(root, ordering)


def _assign_minutes(node: _CategoryNode, minutes_lookup: dict[str, dict[date, int]]) -> None:
    if node.full_name:
        own = minutes_lookup.get(node.full_name.lower(), {})
        node.minutes = dict(own)
    else:
        node.minutes = {}
    for child in node.children:
        _assign_minutes(child, minutes_lookup)


def _write_jf_excel_workbook(root: _CategoryNode, dates: Sequence[date], path: Path) -> None:
    from openpyxl import Workbook  # type: ignore[import]
    from openpyxl.styles import Alignment  # type: ignore[import]
    from openpyxl.utils import get_column_letter  # type: ignore[import]

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Categories"

    headers = ["Category"] + [day.strftime("%A") for day in dates]
    sheet.append(headers)
    date_row = ["Date"] + [day.strftime("%Y-%m-%d") for day in dates]
    sheet.append(date_row)
    sheet.freeze_panes = "B3"

    total_nodes = 0
    top_level_count = len(root.children)
    for index, top_node in enumerate(root.children):
        for node in _walk(top_node):
            total_nodes += 1
            row = [_display_name(node)]
            for day in dates:
                minutes = node.minutes.get(day, 0)
                row.append("" if minutes == 0 else minutes)
            sheet.append(row)
            category_cell = sheet.cell(row=sheet.max_row, column=1)
            indent_level = max(0, node.depth - 1)
            category_cell.alignment = Alignment(indent=indent_level)
        if index < top_level_count - 1:
            sheet.append([""] * len(headers))

    sheet.column_dimensions[get_column_letter(1)].width = 48
    for column_index in range(2, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 14

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["Generated", datetime.utcnow().isoformat(timespec="seconds") + "Z"])
    metadata.append(["Categories", total_nodes])
    metadata.append(["Date columns", len(dates)])

    workbook.save(path)


def _apply_ordering(node: _CategoryNode, ordering: _CategoryOrdering) -> None:
    if not node.children:
        return
    node.children.sort(key=lambda child: (_ordering_key(child, ordering), child.name.lower()))
    for child in node.children:
        _apply_ordering(child, ordering)


def _ordering_key(node: _CategoryNode, ordering: _CategoryOrdering) -> int:
    if not node.full_name:
        return -1
    parts = _split_category_path(node.full_name)
    return ordering.order_for(parts)


def _display_name(node: _CategoryNode) -> str:
    return node.name if node.full_name else "(root)"


def _walk(node: _CategoryNode) -> Iterable[_CategoryNode]:
    yield node
    for child in node.children:
        yield from _walk(child)


def _collect_date_buckets(entries: Sequence[Entry]) -> list[date]:
    if not entries:
        return []
    earliest = min(entry.segment_start.date() for entry in entries)
    latest = max(entry.segment_end.date() for entry in entries)
    span: list[date] = []
    cursor = earliest
    while cursor <= latest:
        span.append(cursor)
        cursor += timedelta(days=1)
    return span
